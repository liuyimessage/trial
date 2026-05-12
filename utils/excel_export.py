"""Excel export utility — generates a colour-coded L2 Business Case workbook."""
import io
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# Brand colours (matching the L2 template)
_DARK_BLUE  = "00003366"
_MID_BLUE   = "004472C4"
_LIGHT_BLUE = "00BDD7EE"
_PALE_BLUE  = "00DEEAF1"
_YELLOW     = "00FFFF00"
_WHITE      = "00FFFFFF"
_LIGHT_GREY = "00F2F2F2"
_GREEN      = "0070AD47"
_ORANGE     = "00FF8C00"

_THIN = Side(style="thin", color="00000000")
_ALL_BORDERS = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _fill(hex_rgb: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_rgb)


def _font(bold=False, size=11, color="00000000", italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")


def _hdr_font() -> Font:
    return Font(bold=True, size=11, color=_WHITE, name="Calibri")


def export_business_case(fields: dict) -> bytes:
    """
    Build a single-sheet L2 Business Case workbook from `fields` dict.
    Returns raw bytes ready for st.download_button.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"_SPRINT_ - {fields.get('initiative_name', 'Business Case')}"

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    def wrow(row, col_a, col_b, col_c="", col_d="",
             fill_a=None, fill_b=None, bold_a=False, bold_b=False,
             border=True, num_fmt_c=None, num_fmt_d=None):
        for ci, (val, fill, bold, num_fmt) in enumerate(zip(
                [col_a, col_b, col_c, col_d],
                [fill_a, fill_b, None, None],
                [bold_a, bold_b, False, False],
                [None, None, num_fmt_c, num_fmt_d],
        ), start=1):
            cell = ws.cell(row=row, column=ci, value=val)
            if fill:
                cell.fill = _fill(fill)
            if bold:
                cell.font = Font(bold=True, name="Calibri", size=11)
            if num_fmt:
                cell.number_format = num_fmt
            if border:
                cell.border = _ALL_BORDERS
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    # ── Title block ────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "L2 Business Case"
    title_cell.fill = _fill(_DARK_BLUE)
    title_cell.font = Font(bold=True, size=16, color=_WHITE, name="Calibri")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:D2")
    sub_cell = ws["A2"]
    sub_cell.value = fields.get("initiative_name", "")
    sub_cell.fill = _fill(_MID_BLUE)
    sub_cell.font = _hdr_font()
    sub_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    # ── Section A — Initiative Identity ────────────────────────────────────────
    r = 3
    wrow(r, "SECTION A — INITIATIVE IDENTITY", "", fill_a=_MID_BLUE, fill_b=_MID_BLUE, bold_a=True)
    ws.row_dimensions[r].height = 18

    pairs_a = [
        ("Initiative Name",     fields.get("initiative_name", "")),
        ("Category",            fields.get("category", "")),
        ("Department / Owner",  fields.get("department", "")),
        ("WAVE ID",             fields.get("wave_id", "")),
        ("Workstream",          fields.get("ws", "")),
        ("Stage",               fields.get("stage", "")),
        ("Status",              fields.get("status", "")),
    ]
    for label, val in pairs_a:
        r += 1
        wrow(r, label, val, fill_a=_PALE_BLUE)
        ws.row_dimensions[r].height = 16

    # ── Section B — Financial Inputs ───────────────────────────────────────────
    r += 1
    wrow(r, "SECTION B — FINANCIAL INPUTS", "", fill_a=_MID_BLUE, fill_b=_MID_BLUE, bold_a=True)
    ws.row_dimensions[r].height = 18

    try:
        baseline = float(fields.get("baseline_value", 0))
    except (TypeError, ValueError):
        baseline = 0.0
    try:
        saving_rate = float(fields.get("saving_rate", 0)) / 100
    except (TypeError, ValueError):
        saving_rate = 0.0
    try:
        rec_benefit = float(fields.get("recurring_benefit", 0))
    except (TypeError, ValueError):
        rec_benefit = 0.0
    try:
        one_time = float(fields.get("one_time_benefit", 0))
    except (TypeError, ValueError):
        one_time = 0.0
    try:
        impl_cost = float(fields.get("implementation_cost", 0))
    except (TypeError, ValueError):
        impl_cost = 0.0

    pairs_b = [
        ("Baseline Value ($)",              baseline),
        ("Saving Rate (%)",                 saving_rate * 100),
        ("Recurring Benefit ($)",           rec_benefit),
        ("One-Time Benefit ($)",            one_time),
        ("Implementation Cost ($)",         impl_cost),
        ("Net Benefit ($)",                 rec_benefit + one_time - impl_cost),
    ]
    num_fmts = ['#,##0.00', '0.00"%"', '#,##0', '#,##0', '#,##0', '#,##0']
    for (label, val), nf in zip(pairs_b, num_fmts):
        r += 1
        row_cells = [ws.cell(row=r, column=1, value=label),
                     ws.cell(row=r, column=2, value=val)]
        row_cells[0].fill = _fill(_LIGHT_BLUE)
        row_cells[1].fill = _fill(_WHITE)
        row_cells[1].number_format = nf
        for c in row_cells:
            c.border = _ALL_BORDERS
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 16

    # Mark C11 and C15 for baseline & recurring benefit (template reference cells)
    ws["C11"].value = baseline
    ws["C11"].number_format = "#,##0"
    ws["C15"].value = rec_benefit
    ws["C15"].number_format = "#,##0"

    # ── Section C — Execution Plan ─────────────────────────────────────────────
    r += 1
    wrow(r, "SECTION C — EXECUTION PLAN", "", fill_a=_MID_BLUE, fill_b=_MID_BLUE, bold_a=True)
    ws.row_dimensions[r].height = 18

    exec_steps = fields.get("execution_steps", [])
    if isinstance(exec_steps, str):
        exec_steps = [s.strip() for s in exec_steps.split("\n") if s.strip()]
    if not exec_steps:
        exec_steps = ["(No steps provided)"]

    for i, step in enumerate(exec_steps, 1):
        r += 1
        wrow(r, f"Step {i}", step, fill_a=_LIGHT_GREY)
        ws.row_dimensions[r].height = 16

    # ── Section D — Assumptions ────────────────────────────────────────────────
    r += 1
    wrow(r, "SECTION D — ASSUMPTIONS & RISKS", "", fill_a=_MID_BLUE, fill_b=_MID_BLUE, bold_a=True)
    ws.row_dimensions[r].height = 18

    assumptions = fields.get("assumptions", "")
    if not assumptions:
        assumptions = "(No assumptions provided)"
    r += 1
    ws.merge_cells(f"A{r}:B{r}")
    cell = ws[f"A{r}"]
    cell.value = assumptions
    cell.fill = _fill(_WHITE)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.border = _ALL_BORDERS
    ws.row_dimensions[r].height = max(30, len(assumptions) // 3)

    # ── Freeze panes & print settings ─────────────────────────────────────────
    ws.freeze_panes = "A3"
    ws.print_title_rows = "1:2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
